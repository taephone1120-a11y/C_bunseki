import io
import re
import time
import random
import requests
import json
import textwrap
import math
import threading
import traceback
from collections import Counter
from types import SimpleNamespace
import numpy as np
import pandas as pd
import streamlit as st
from bs4 import BeautifulSoup
from urllib.parse import quote, urljoin, urlparse, parse_qs, urlencode, urlunparse
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

# =============================================
#   デザインとヘッダー設定
# =============================================
st.set_page_config(page_title="Creema市場リサーチツール", page_icon="💎", layout="wide")

st.markdown("""
    <style>
    .block-container { padding-top: 2.5rem !important; padding-bottom: 2rem !important; }
    html, body, [data-testid="stMarkdownContainer"] p, .stMarkdown p {
        font-size: 14px !important;
        font-family: "Meiryo", "Helvetica Neue", Arial, sans-serif;
        line-height: 1.5 !important;
    }
    h1 { font-size: 28px !important; font-weight: 700 !important; color: #111111 !important; margin: 0 !important; }
    div[data-testid="stSidebarUserContent"] { padding-top: 1rem !important; }
    div[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h5 { font-size: 13.5px !important; margin-top: 12px !important; margin-bottom: 5px !important; color: #111111 !important; font-weight: 600 !important; }
    
    .stTextInput input, .stNumberInput input, .stDateInput input, div[data-testid="stSelectbox"] div { 
        padding: 4px 8px !important; 
        min-height: 32px !important; 
        height: 32px !important; 
        font-size: 13px !important; 
    }
    
    div[data-testid="stNumberInput"] button {
        display: none !important;
    }
    div[data-testid="stNumberInput"] div[data-baseline="true"] {
        padding-right: 8px !important;
    }
    
    .metric-card {
        background-color: #f8f9fa;
        border-left: 5px solid #1f497d;
        padding: 15px;
        border-radius: 4px;
        margin-bottom: 15px;
    }
    </style>
""", unsafe_allow_html=True)

st.title("💎 Creema市場リサーチツール")
st.markdown('<hr style="border: none; border-top: 1px solid #e6e6e6; margin-top: 15px; margin-bottom: 25px; padding: 0;">', unsafe_allow_html=True)

# =============================================
#   サイドバー：設定エリア ＆ フィルターエリア
# =============================================
st.sidebar.header("⚙️ 取得条件設定")
mode = st.sidebar.radio("収集モードを選択してください", ("キーワード検索", "一覧URL直貼り"))

search_keyword = ""
target_url = ""

if mode == "キーワード検索":
    search_keyword = st.sidebar.text_input("🔍 検索キーワードを入力", value="")
    encoded_keyword = quote(search_keyword)
    target_url = f"https://www.creema.jp/listing?q={encoded_keyword}&active=pc_listing-form"
else:
    target_url = st.sidebar.text_input("🔗 Creemaの一覧URLを入力", value="")

# =============================================
#   🛡️ Creemaサーバーに迷惑をかけないための全体安全設定
# =============================================
# ここは「1人のユーザーの設定」ではなく、このアプリを使う
# 全員・全セッションに共通で効く「アプリ全体の上限」です。
# st.cache_resource を使うことで、再実行やセッションをまたいでも
# プロセス内に1つだけ残る「本当のグローバル状態」として扱えます。
# （※複数サーバー・複数プロセスに分散デプロイする場合はこの限りではなく、
# 　その場合は外部のRedis等で共有状態を持つ必要があります）

HARD_MAX_ITEMS = 300               # 1回のリサーチで取得できる件数の絶対上限（1000→300に縮小）
GLOBAL_MIN_REQUEST_INTERVAL = 0.5  # 全ユーザー合算で、Creemaへのリクエスト最小間隔（秒）

# --- 全ユーザー・全セッションで共有されるペーシング用ロック ---
# 同時に何人がボタンを押しても、実際にCreemaへ送信されるリクエストは
# 全員このロック1本の後ろに並び、0.5秒に1回以下のペースに間引かれる。
# そのため「同時実行を人数で制限する」仕組みは不要（かつ、カウンターが
# ズレて固まった場合に自分自身を締め出すリスクがあるだけなので撤去した）。
@st.cache_resource
def get_shared_state():
    return {
        "pace_lock": threading.Lock(),
        "last_request_ts": 0.0,
    }

_shared = get_shared_state()

def global_pace_wait():
    with _shared["pace_lock"]:
        now = time.monotonic()
        wait = GLOBAL_MIN_REQUEST_INTERVAL - (now - _shared["last_request_ts"])
        if wait > 0:
            time.sleep(wait)
        _shared["last_request_ts"] = time.monotonic()

max_items = st.sidebar.number_input(
    "🔢 取得する商品件数",
    min_value=1,
    max_value=HARD_MAX_ITEMS,
    value=min(100, HARD_MAX_ITEMS),
    step=10,
    help=f"Creemaサーバーへの負荷を抑えるため、1回あたり最大{HARD_MAX_ITEMS}件までに制限しています。"
)
max_items = min(int(max_items), HARD_MAX_ITEMS)

# 取得速度は「高速（おすすめ）」に固定（画面には表示しない）
speed_mode = "高速（おすすめ）"

SPEED_CONFIG = {
    "高速（おすすめ）": {"review_pages": 20, "workers_large": 6, "workers_small": 8, "sleep": 0.12},
}
CURRENT_SPEED = SPEED_CONFIG[speed_mode]

st.sidebar.markdown('---')
st.sidebar.header("📅 評価日データの取得")
include_eval_dates = st.sidebar.checkbox(
    "評価日1〜3を取得する",
    value=True,
    help="OFFにすると、この列は結果に含まれません。"
)
include_first_review_date = st.sidebar.checkbox(
    "作家の一番初めの評価日を取得する",
    value=True,
    help="OFFにすると、商品1件ごとの追加リクエストが1回減るため、リサーチが少し速くなります。"
)

# 診断ログと通信リトライ設定は内部処理用に残し、画面には表示しない
show_diagnostics = False
skip_line_notify = False
retry_count = 4
retry_base_wait = 2.0

RETRY_STATUS_CODES = {403, 408, 425, 429, 500, 502, 503, 504, 520, 521, 522, 523, 524}

start_button = st.sidebar.button("🚀 リサーチを開始する", type="primary")

# フィルターエリア
st.sidebar.markdown('---')
st.sidebar.header("📊 表示データの絞り込み")

# 十分に広い日付範囲を定義
past_limit_date = datetime.strptime("2000-01-01", "%Y-%m-%d").date()
future_limit_date = datetime.strptime("2030-12-31", "%Y-%m-%d").date()

# 1. 価格
st.sidebar.markdown("**価格(円)**")
col_price1, col_price_tilde, col_price2 = st.sidebar.columns([4, 1, 4])
with col_price1:
    min_price = st.number_input("価格（最小）", min_value=0, value=0, label_visibility="collapsed")
with col_price_tilde:
    st.markdown("<div style='text-align: center; line-height: 32px;'>〜</div>", unsafe_allow_html=True)
with col_price2:
    max_price = st.number_input("価格（最大）", min_value=0, value=99999, label_visibility="collapsed")

# 3. その商品の直近1ヶ月の評価数
st.sidebar.markdown("**その商品の直近1ヶ月の評価数**")
col_recent1, col_recent_tilde, col_recent2 = st.sidebar.columns([4, 1, 4])
with col_recent1:
    min_recent_review = st.number_input(
        "その商品の直近1ヶ月の評価数（最小）",
        min_value=0,
        value=0,
        label_visibility="collapsed"
    )
with col_recent_tilde:
    st.markdown("<div style='text-align: center; line-height: 32px;'>〜</div>", unsafe_allow_html=True)
with col_recent2:
    max_recent_review = st.number_input(
        "その商品の直近1ヶ月の評価数（最大）",
        min_value=0,
        value=99999,
        label_visibility="collapsed"
    )

# 2. 作家の総評価数

st.sidebar.markdown("**作家の総評価数**")
col_rev1, col_rev_tilde, col_rev2 = st.sidebar.columns([4, 1, 4])
with col_rev1:
    min_rev = st.number_input("作家の総評価数（最小）", min_value=0, value=0, label_visibility="collapsed")
with col_rev_tilde:
    st.markdown("<div style='text-align: center; line-height: 32px;'>〜</div>", unsafe_allow_html=True)
with col_rev2:
    max_rev = st.number_input("作家の総評価数（最大）", min_value=0, value=99999, label_visibility="collapsed")


# =============================================
#   高速通信用：Session使い回し
# =============================================
_thread_local = threading.local()
_cache_lock = threading.Lock()
_rating_page_cache = {}

# =============================================
#   🩺 診断ログ：止まった原因を見える化
# =============================================
_diag_lock = threading.Lock()
_diag_logs = []
_diag_stats = Counter()

def reset_diagnostics():
    global _diag_logs, _diag_stats
    with _diag_lock:
        _diag_logs = []
        _diag_stats = Counter()

def diag_log(stage, message, url="", level="INFO"):
    """画面表示用・ターミナル確認用の診断ログ。スレッドからも呼べるようにロックする。"""
    entry = {
        "時刻": datetime.now().strftime("%H:%M:%S"),
        "レベル": level,
        "工程": stage,
        "内容": str(message),
        "URL": str(url)[:300],
    }
    with _diag_lock:
        _diag_logs.append(entry)
        if len(_diag_logs) > 1000:
            del _diag_logs[:-1000]
        _diag_stats[f"{level}:{stage}"] += 1
    print(f"[{entry['時刻']}] [{level}] {stage}: {message} {url}")

def diag_count(key, amount=1):
    with _diag_lock:
        _diag_stats[key] += amount

def get_diagnostics():
    with _diag_lock:
        return list(_diag_logs), dict(_diag_stats)

def get_fast_session():
    """スレッドごとにrequests.Sessionを使い回して、接続確立の時間を減らす。"""
    if not hasattr(_thread_local, "session"):
        session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(pool_connections=50, pool_maxsize=50)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        _thread_local.session = session
    return _thread_local.session

def fast_get(url, headers=None, timeout=10, retry_label="通信"):
    """
    Creema側に一時的に弾かれた/混み合った時の再アクセス付きGET。
    対象: 403, 429, 500系, タイムアウトなど。
    404など恒久的な失敗は基本リトライしない。
    """
    session = get_fast_session()
    last_error = None
    max_attempts = int(retry_count) + 1

    for attempt in range(1, max_attempts + 1):
        try:
            # 🛡️ 全ユーザー・全スレッド共通の間隔を必ず空けてから送信する
            global_pace_wait()
            res = session.get(url, headers=headers, timeout=timeout)

            # 成功、またはリトライ対象外のHTTPステータスならそのまま返す
            if res.status_code == 200 or res.status_code not in RETRY_STATUS_CODES:
                if attempt > 1 and res.status_code == 200:
                    diag_count("リトライ成功")
                    diag_log(retry_label, f"リトライ成功 attempt={attempt} status=200", url, "INFO")
                elif attempt > 1:
                    diag_count(f"リトライ後_HTTP_{res.status_code}")
                    diag_log(retry_label, f"リトライ後も対象外status={res.status_code} attempt={attempt}", url, "WARN")
                return res

            # リトライ対象のHTTPステータス
            last_error = f"HTTP {res.status_code}"
            diag_count(f"HTTP_{res.status_code}")

            if attempt < max_attempts:
                wait_sec = float(retry_base_wait) * attempt + random.uniform(0.2, 1.2)
                diag_count("リトライ実行")
                diag_log(
                    retry_label,
                    f"弾かれ/一時失敗 status={res.status_code} → {wait_sec:.1f}秒後に再アクセス attempt={attempt}/{max_attempts}",
                    url,
                    "WARN"
                )
                time.sleep(wait_sec)
                continue

            diag_count("リトライ上限到達")
            diag_log(retry_label, f"リトライ上限到達 status={res.status_code} attempts={max_attempts}", url, "ERROR")
            return res

        except requests.exceptions.RequestException as e:
            last_error = f"{type(e).__name__}: {e}"
            diag_count(f"通信例外_{type(e).__name__}")

            if attempt < max_attempts:
                wait_sec = float(retry_base_wait) * attempt + random.uniform(0.2, 1.2)
                diag_count("リトライ実行")
                diag_log(
                    retry_label,
                    f"通信例外 {type(e).__name__} → {wait_sec:.1f}秒後に再アクセス attempt={attempt}/{max_attempts}",
                    url,
                    "WARN"
                )
                time.sleep(wait_sec)
                continue

            diag_count("リトライ上限到達")
            diag_log(retry_label, f"通信例外でリトライ上限到達: {last_error}", url, "ERROR")
            raise

    # 通常ここには来ないが、型安全のため最後に1回アクセス
    global_pace_wait()
    return session.get(url, headers=headers, timeout=timeout)

def cached_fast_get(url, headers=None, timeout=10, force_refresh=False):
    """レビューページなど、同じURLを何度も読む可能性があるページをメモリに保存する。
    force_refresh=True の場合はキャッシュを使わず、もう一度取得し直す。
    """
    if not force_refresh:
        with _cache_lock:
            cached = _rating_page_cache.get(url)

        if cached is not None:
            status_code, final_url, content = cached
            return SimpleNamespace(status_code=status_code, url=final_url, content=content)

    res = fast_get(url, headers=headers, timeout=timeout, retry_label="レビューページ")

    if res.status_code == 200:
        with _cache_lock:
            _rating_page_cache[url] = (res.status_code, res.url, res.content)

    return res

# =============================================
#   📲 LINE通知関数
# =============================================
def send_line_notification(keyword_or_url, item_count):
    LINE_ACCESS_TOKEN = "SsJj64qF912H/fusrwNgsiMS6bgJqv5C9i5Rx1HlHAmux8AmFlC7Q9Pnx5pbQD/4LXbi2ftiFf1zalCCDcGQAcXBxfakpnkBPLZkKzn5G2gbuQc2vkcn2GbCJ2Yf1HmfEWQoo8KbqqJn4/tsoPr4TwdB04t89/1O/w1cDnyilFU="
    LINE_USER_ID = "Ub5228833332f8fd37bbd3d9072853f2c"
    url = "https://api.line.me/v2/bot/message/push"
    headers = { "Content-Type": "application/json", "Authorization": f"Bearer {LINE_ACCESS_TOKEN}" }
    message_text = f"💎 【Creemaツール】利用通知\n\nリサーチ開始！\n内容:\n{keyword_or_url}\n上限: {item_count} 件"
    try: requests.post(url, headers=headers, json={"to": LINE_USER_ID, "messages": [{"type": "text", "text": message_text}]}, timeout=5)
    except: pass

# =============================================
#   Excelダウンロード用バイナリ生成関数
# =============================================
def convert_df_to_excel(df):
    export_df = df.copy()

    def remove_illegal_chars(val):
        if isinstance(val, str):
            cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", val)
            return "".join(ch for ch in cleaned if ch.isprintable() or ch in "\n\r\t")
        return val

    for col in export_df.columns:
        export_df[col] = export_df[col].apply(remove_illegal_chars)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="リサーチ結果", index=False)
        worksheet = writer.sheets["リサーチ結果"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_font = Font(name="Meiryo", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        data_font = Font(name="Meiryo", size=10)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        # セル装飾
        for row in worksheet.iter_rows(min_row=1, max_row=len(export_df) + 1):
            for cell in row:
                cell.border = thin_border

                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.font = data_font

                    # 商品URL列をリンク化
                    if cell.column == 5 and cell.value and str(cell.value).startswith("http"):
                        cell.hyperlink = cell.value
                        cell.font = Font(name="Meiryo", size=10, color="0563C1", underline="single")

                    # 基本の配置
                    if cell.column in [1, 4, 6]:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif cell.column in [7, 8, 9, 10, 11, 12, 13]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 列幅調整
        for col in worksheet.columns:
            col_letter = col[0].column_letter
            header_name = col[0].value
            max_len = max(len(str(cell.value or "")) for cell in col)

            # 商品名は少しきゅっと狭める
            if header_name == "商品名":
                worksheet.column_dimensions[col_letter].width = 28
                for cell in col:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="top",
                        wrap_text=True
                    )

            # 作品紹介文は広め＋折り返し
            elif header_name == "作品紹介文":
                worksheet.column_dimensions[col_letter].width = 45
                for cell in col:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="top",
                        wrap_text=True
                    )

            # 商品URLは短め
            elif header_name == "商品URL":
                worksheet.column_dimensions[col_letter].width = 18

            # その他の列
            else:
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 25)

    return output.getvalue()

# =============================================
#   レビューページ解析補助
# =============================================
def extract_review_blocks_from_soup(soup):
    """Creemaのレビューブロック候補を複数セレクタで拾う。HTML変更に少し強くする。"""
    selectors = [
        ".p-creator-rating-list__item",
        ".p-creator-rating-rating__content",
        "li[class*=rating]",
        "div[class*=rating][class*=item]",
        "div[class*=review][class*=item]",
    ]
    for selector in selectors:
        blocks = soup.select(selector)
        # 日付が含まれるブロックだけを優先
        dated_blocks = [b for b in blocks if re.search(r"\d{4}\.\d{2}\.\d{2}", b.get_text(" ", strip=True))]
        if dated_blocks:
            return dated_blocks, selector
        if blocks:
            return blocks, selector
    return [], ""

def looks_like_blocked_or_unexpected_page(soup):
    """200で返ってきても、アクセス制限・ログイン誘導・想定外ページっぽい場合を判定する。"""
    text = soup.get_text(" ", strip=True)[:3000]
    suspicious_words = [
        "アクセスが集中", "しばらく時間をおいて", "ただいま混み合って",
        "Forbidden", "Too Many Requests", "Access Denied", "認証", "ログイン",
        "エラーが発生", "ページが見つかりません"
    ]
    return any(word in text for word in suspicious_words)

# =============================================
#   1件詳細解析用パーツ
# =============================================
def _internal_fetch_item(item_data, headers, one_month_ago, include_eval_dates=True, include_first_review_date=True):
    link = item_data["link"]
    creator = item_data["creator"]
    title = item_data["title"]
    price = item_data["price"]

    favorite = 0
    purchase_display = "0人"
    review = 0  # クリエイター欄の作家の総評価数
    recent_review_display = "0件"
    first_review_date = "データなし"
    description_text = "取得失敗"
    recent_sales = ["ー", "ー", "ー"]

    # 評価日は3ヶ月以内を対象にする
    three_months_ago = datetime.now() - timedelta(days=90)

    try:
        res = fast_get(link, headers=headers, timeout=10, retry_label="詳細ページ")
        if res.status_code != 200:
            diag_count("詳細ページ_HTTP失敗")
            diag_count(f"詳細ページ_HTTP_{res.status_code}")
            diag_log("詳細ページ", f"商品詳細ページの取得に失敗 status={res.status_code} ※リトライ後", link, "ERROR")
            return None

        soup = BeautifulSoup(res.content, "html.parser")

        # =========================
        # 作品紹介文取得
        # =========================
        desc_tag = soup.select_one(
            "#introduction .p-item-detail-description, "
            ".p-item-detail-description.js-internal-link, "
            ".p-item-detail__description, "
            ".p-item-detail-body__description"
        )

        if desc_tag:
            def extract_description_with_br(tag):
                parts = []

                def walk(node):
                    # テキストノード
                    if isinstance(node, str):
                        text = node

                        # HTMLのインデントだけの空白は無視
                        if text.strip() == "":
                            return

                        # 行頭行末の余計な空白だけ削る
                        parts.append(text.strip())
                        return

                    # brタグは1個につき改行1つ
                    if getattr(node, "name", None) == "br":
                        parts.append("\n")
                        return

                    # aタグは表示テキストを残す
                    if getattr(node, "name", None) == "a":
                        text = node.get_text(strip=True)
                        href = node.get("href", "")

                        if href and href.startswith("/"):
                            href = "https://www.creema.jp" + href

                        if text:
                            parts.append(text)
                        elif href:
                            parts.append(href)
                        return

                    # その他のタグは中身を再帰的に見る
                    if hasattr(node, "children"):
                        for child in node.children:
                            walk(child)

                for child in tag.children:
                    walk(child)

                text = "".join(parts)

                # 改行コードを統一
                text = text.replace("\r\n", "\n").replace("\r", "\n")

                # 各行の前後スペースだけ削る
                lines = [line.strip() for line in text.split("\n")]
                text = "\n".join(lines)

                # 先頭・末尾の余計な改行だけ削除
                text = text.strip("\n ")

                return text

            description_text = extract_description_with_br(desc_tag)

        else:
            description_text = "取得失敗"

        # =========================
        # クリエイターの作家の総評価数取得
        # =========================
        creator_rating_tag = soup.select_one(
            "#js-creator-rating-average .p-item-detail-creator__rating-count a"
        )

        if creator_rating_tag:
            rating_text = creator_rating_tag.get_text(strip=True)
            rating_num = re.sub(r"\D", "", rating_text)
            review = int(rating_num) if rating_num else 0
        else:
            rating_count_area = soup.select_one(
                "#js-creator-rating-average .p-item-detail-creator__rating-count"
            )
            if rating_count_area:
                rating_text = rating_count_area.get_text(strip=True)
                rating_num = re.sub(r"\D", "", rating_text)
                review = int(rating_num) if rating_num else 0

        # =========================
        # お気に入り数取得
        # =========================
        fav_btn = soup.find(
            lambda tag: tag.name in ["button", "a"] and "お気に入りに追加" in tag.text
        )

        if fav_btn:
            num_part = fav_btn.select_one(".js-like-item-number, b, span")
            fav_text = re.sub(r"\D", "", num_part.text if num_part else fav_btn.text)
            favorite = int(fav_text) if fav_text else 0

        # =========================
        # 購入者数取得
        # =========================
        buy_container = soup.select_one(".p-item-detail-info__item--left")

        if buy_container:
            text = buy_container.get_text(strip=True)

            if "10人以上購入" in text:
                purchase_display = "10人以上"
            else:
                match = re.search(r"(\d+)人購入", text)
                if match:
                    purchase_display = f"{match.group(1)}人"

        # =========================
        # レビューページ解析
        # =========================
        rating_link_tag = soup.select_one('a[href*="/rating/sale"]')

        if rating_link_tag:
            href_attr = rating_link_tag["href"]
            base_rating_url = href_attr if href_attr.startswith("http") else "https://www.creema.jp" + href_attr

            if "?" in base_rating_url:
                base_rating_url = base_rating_url.split("?")[0]

            # /creator/数字/rating/sale のまま page=2 を付けると、
            # page が消えることがあるので、先に正規URLへ変換する
            try:
                canonical_res = fast_get(base_rating_url, headers=headers, timeout=10, retry_label="レビューページ正規URL")
                if canonical_res.status_code == 200:
                    canonical_rating_url = canonical_res.url.split("?")[0]
                else:
                    canonical_rating_url = base_rating_url
            except Exception:
                canonical_rating_url = base_rating_url

            all_found_dates = []
            seen_review_keys = set()

            # 商品名だけで一致判定
            target_name = "".join(title.split())

            print(f"[{title[:15]}...] レビュー探索開始")
            print("元の評価URL:", base_rating_url)
            print("正規評価URL:", canonical_rating_url)

            # =========================
            # 対象商品のレビュー日を取得
            # =========================
            # その商品の直近1ヶ月の評価数：
            #   対象商品のレビューのうち、直近30日以内の件数
            #
            # 評価日1〜3：
            #   レビュー日付が3ヶ月より古くなるページまで見て、
            #   その中で見つかった対象商品の直近3回分の日付を入れる
            # =========================
            max_review_pages = CURRENT_SPEED["review_pages"]

            for current_page in range(1, max_review_pages + 1):
                page_url = f"{canonical_rating_url}?page={current_page}"
                print(f" - 取得URL: {page_url}")

                r_res = cached_fast_get(page_url, headers=headers, timeout=10)
                print(" - 最終URL:", r_res.url)

                if r_res.status_code != 200:
                    diag_count("レビューページ_HTTP失敗")
                    diag_log("レビューページ", f"{current_page}ページ目 status={r_res.status_code} のためこの商品のレビュー探索を終了", page_url, "WARN")
                    print(f" - {current_page}ページ目: ステータスコード {r_res.status_code} のため終了")
                    break

                r_soup = BeautifulSoup(r_res.content, "html.parser")
                blocks, block_selector = extract_review_blocks_from_soup(r_soup)

                # status=200でも、空ページ・アクセス制限・HTML変更っぽい時は、
                # キャッシュを使わず1回だけ取り直す。
                if not blocks:
                    diag_count("レビューブロック0件_初回")
                    if looks_like_blocked_or_unexpected_page(r_soup):
                        diag_count("レビューページ_想定外HTML")
                        diag_log("レビューページ", f"{current_page}ページ目が想定外HTML/制限ページっぽいため、キャッシュなしで再取得", page_url, "WARN")
                    else:
                        diag_log("レビューページ", f"{current_page}ページ目でレビューブロック0件のため、キャッシュなしで再取得", page_url, "WARN")

                    r_res_retry = cached_fast_get(page_url, headers=headers, timeout=12, force_refresh=True)
                    if r_res_retry.status_code == 200:
                        r_soup_retry = BeautifulSoup(r_res_retry.content, "html.parser")
                        retry_blocks, retry_selector = extract_review_blocks_from_soup(r_soup_retry)
                        if retry_blocks:
                            blocks = retry_blocks
                            block_selector = retry_selector
                            r_soup = r_soup_retry
                            diag_count("レビューブロック0件_再取得成功")
                            diag_log("レビューページ", f"再取得でレビューブロック取得成功 selector={block_selector} 件数={len(blocks)}", page_url, "INFO")
                        else:
                            diag_count("レビューブロック0件")
                            diag_log("レビューページ", f"再取得後もレビューブロック0件。レビューなし/HTML変更/対象外ページの可能性", page_url, "WARN")
                            print(f" - {current_page}ページ目: 再取得後もレビューブロックが0件のため終了")
                            break
                    else:
                        diag_count("レビューページ_HTTP失敗_再取得")
                        diag_log("レビューページ", f"レビューブロック0件後の再取得で status={r_res_retry.status_code}", page_url, "WARN")
                        break

                print(f" - {current_page}ページ目: レビューブロック {len(blocks)}件 selector={block_selector}")

                found_in_page = 0
                page_dates = []

                for block in blocks:
                    # 日付取得
                    date_tag = block.select_one(
                        ".p-creator-rating-rating__date, "
                        ".p-creator-rating-list__item-date"
                    )

                    if date_tag:
                        date_text = date_tag.get_text(" ", strip=True)
                    else:
                        date_text = block.get_text(" ", strip=True)

                    d_match = re.search(r"(\d{4})\.(\d{2})\.(\d{2})", date_text)

                    if not d_match:
                        continue

                    found_date = datetime(
                        int(d_match.group(1)),
                        int(d_match.group(2)),
                        int(d_match.group(3))
                    )

                    # ページ全体の日付。探索終了判定に使う
                    page_dates.append(found_date)

                    # 商品名リンク取得
                    item_name_tag = block.select_one(
                        '.p-creator-rating-rating__title a[href*="/item/"], '
                        '.p-creator-rating-list__item-title a[href*="/item/"]'
                    )

                    if not item_name_tag:
                        item_links = block.select('a[href*="/item/"]')
                        for a in item_links:
                            if a.get_text(strip=True):
                                item_name_tag = a
                                break

                    if not item_name_tag:
                        continue

                    review_href = item_name_tag.get("href", "")
                    review_item_name = "".join(item_name_tag.get_text(strip=True).split())

                    is_same_by_name = (
                        target_name in review_item_name
                        or review_item_name in target_name
                    )

                    if not is_same_by_name:
                        continue

                    print("【一致】対象商品として処理")
                    print("対象商品名:", target_name[:80])
                    print("レビュー商品名:", review_item_name[:80])
                    print("評価日:", found_date.strftime("%Y.%m.%d"))
                    print("-" * 50)

                    # 二重取得防止
                    review_text = block.get_text(" ", strip=True)
                    review_key = f"{review_href}_{found_date.strftime('%Y.%m.%d')}_{review_text[:100]}"

                    if review_key in seen_review_keys:
                        continue

                    seen_review_keys.add(review_key)
                    all_found_dates.append(found_date)
                    found_in_page += 1

                print(f" - {current_page}ページ目終了: 対象商品一致 {found_in_page}件")

                # このページの日付がすべて3ヶ月より古いなら終了
                if page_dates:
                    newest_date_in_page = max(page_dates)
                    oldest_date_in_page = min(page_dates)

                    print(
                        f" - {current_page}ページ目の日付範囲: "
                        f"{oldest_date_in_page.strftime('%Y.%m.%d')} 〜 "
                        f"{newest_date_in_page.strftime('%Y.%m.%d')}"
                    )

                    if newest_date_in_page < three_months_ago:
                        print(" - このページはすべて3ヶ月より古いため、探索終了")
                        break

                time.sleep(CURRENT_SPEED["sleep"])

            # =========================
            # 評価日1〜3
            # =========================
            all_found_dates.sort(reverse=True)

            # 3ヶ月以内の対象商品レビューだけを評価日として使う
            recent_three_month_dates = [
                d for d in all_found_dates
                if d >= three_months_ago
            ]

            # 購入者数を数値化する
            # 「0人」→ 0
            # 「1人」→ 1
            # 「10人以上」→ 10
            def parse_purchase_num(text):
                if not isinstance(text, str):
                    return 0

                if "10人以上" in text:
                    return 10

                m = re.search(r"(\d+)", text)
                return int(m.group(1)) if m else 0

            purchase_num = parse_purchase_num(purchase_display)

            # =========================
            # 評価日1〜3の表示ルール
            # =========================
            # 購入者数0人：
            #   すべて「ー」
            #
            # 購入者数1人：
            #   評価日1だけ表示対象。2〜3は「ー」
            #
            # 購入者数2人：
            #   評価日1〜2だけ表示対象。3は「ー」
            #
            # 購入者数3人以上：
            #   評価日1〜3すべて表示対象
            #
            # 表示対象の欄について：
            #   3ヶ月以内の日付が取れた欄は日付
            #   日付が取れなかった欄は「3ヶ月以上前」
            # =========================

            recent_sales = ["ー", "ー", "ー"]

            if include_eval_dates:
                display_slots = min(purchase_num, 3)

                for i in range(display_slots):
                    if i < len(recent_three_month_dates):
                        recent_sales[i] = recent_three_month_dates[i].strftime("%Y.%m.%d")
                    else:
                        recent_sales[i] = "3ヶ月以上前"
            else:
                recent_sales = ["対象外(OFF)", "対象外(OFF)", "対象外(OFF)"]

            # =========================
            # 対象商品のその商品の直近1ヶ月の評価数
            # =========================
            recent_month_count = sum(
                1 for d in all_found_dates
                if d >= one_month_ago
            )
            recent_review_display = f"{recent_month_count}件"

            # =========================
            # 作家の一番初めの評価日
            # =========================
            # クリエイターの作家の総評価数から最終ページを計算し、
            # 最終ページ内の日付の中で一番古い日付を使う
            # OFFの場合は追加リクエストをかけずにスキップする
            if include_first_review_date:
                try:
                    if review > 0:
                        last_page = math.ceil(review / 20)

                        last_page_url = f"{canonical_rating_url}?page={last_page}"
                        print("作家の一番初めの評価日チェックURL:", last_page_url)

                        last_res = cached_fast_get(last_page_url, headers=headers, timeout=10)
                        print("作家の一番初めの評価日チェック 最終URL:", last_res.url)

                        if last_res.status_code == 200:
                            last_soup = BeautifulSoup(last_res.content, "html.parser")
                            last_page_text = last_soup.get_text(" ", strip=True)

                            date_matches = re.findall(
                                r"\(\s*(\d{4}\.\d{2}\.\d{2})\s*\)",
                                last_page_text
                            )

                            if date_matches:
                                oldest_date = min(
                                    datetime.strptime(d, "%Y.%m.%d")
                                    for d in date_matches
                                )
                                first_review_date = oldest_date.strftime("%Y.%m.%d")

                except Exception as e:
                    diag_count("初回評価日_取得失敗")
                    diag_log("初回評価日", f"取得エラー: {type(e).__name__}: {e}", canonical_rating_url, "WARN")
                    print("作家の一番初めの評価日取得エラー:", e)
            else:
                first_review_date = "対象外(OFF)"

        else:
            diag_count("レビューページリンクなし")
            diag_log("レビューページ", "商品詳細内に /rating/sale のリンクが見つかりません", link, "WARN")

        diag_count("詳細解析_成功")
        return {
            "No.": 0,
            "商品URL": link,
            "作家名": creator,
            "商品名": title,
            "価格(円)": price,
            "作品紹介文": description_text,
            "その商品の直近1ヶ月の評価数": recent_review_display,
            "作家の総評価数": review,
            "お気に入り数": favorite,
            "購入者数": purchase_display,
            "評価日1": recent_sales[0],
            "評価日2": recent_sales[1],
            "評価日3": recent_sales[2],
            "作家の一番初めの評価日": first_review_date,
           
        }

    except Exception as e:
        diag_count("詳細解析_例外")
        diag_log("詳細解析", f"例外: {type(e).__name__}: {e}", link, "ERROR")
        print("詳細解析エラー:", e)
        print(traceback.format_exc())
        return None



def make_absolute_url(href, base_url="https://www.creema.jp"):
    """相対URLを絶対URLにする。"""
    if not href:
        return ""
    return urljoin(base_url, href)




def normalize_creema_item_url(url):
    """Creemaの商品URLを重複判定用に正規化する。
    例:
    https://www.creema.jp/item/20646858/detail?vkey=...
    → https://www.creema.jp/item/20646858/detail
    """
    if not url:
        return ""

    abs_url = make_absolute_url(url)
    parsed = urlparse(abs_url)

    # /item/数字/detail を取り出す
    m = re.search(r"/item/(\d+)/detail", parsed.path)
    if m:
        return f"https://www.creema.jp/item/{m.group(1)}/detail"

    # 万一 /item/数字 までしか取れない場合も拾う
    m = re.search(r"/item/(\d+)", parsed.path)
    if m:
        return f"https://www.creema.jp/item/{m.group(1)}/detail"

    # item URL以外は、クエリとフラグメントを落とす
    return urlunparse((parsed.scheme or "https", parsed.netloc or "www.creema.jp", parsed.path, "", "", ""))


def set_query_param(url, key, value):
    """URLのクエリに key=value を安全に入れる。"""
    parsed = urlparse(url)
    query = parse_qs(parsed.query, keep_blank_values=True)
    query[key] = [str(value)]
    new_query = urlencode(query, doseq=True)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))


def find_next_listing_url(soup, current_url, page_count):
    """
    Creemaの次ページURLをできるだけ粘って探す。
    以前は a.c-pagination__next だけだったため、HTML差分で1ページ目=10件で止まりやすかった。
    """
    # 1) いちばん従来に近い next ボタン
    selectors = [
        'a.c-pagination__next[href]',
        'a[rel="next"][href]',
        '.c-pagination a[aria-label*="次"][href]',
        '.c-pagination a[href*="page="]',
        'nav a[aria-label*="次"][href]',
        'a[aria-label*="次"][href]',
        'a:contains("次")',  # SoupSieveで使えない環境もあるのでtry側で吸収
    ]

    for selector in selectors:
        try:
            candidates = soup.select(selector)
        except Exception:
            candidates = []
        for a in candidates:
            href = a.get("href", "")
            text = a.get_text(" ", strip=True)
            cls = " ".join(a.get("class", []))
            aria = a.get("aria-label", "")
            if not href:
                continue
            # 明らかな前ページや現在ページは除外
            if "prev" in cls.lower() or "前" in text or "前" in aria:
                continue
            # c-pagination内の page= は複数候補があるので、現在より大きいpageを優先
            abs_url = make_absolute_url(href, current_url)
            parsed = urlparse(abs_url)
            qs = parse_qs(parsed.query)
            page_vals = qs.get("page") or qs.get("p")
            if page_vals:
                try:
                    if int(page_vals[0]) <= page_count:
                        continue
                except Exception:
                    pass
            # nextっぽいものを返す
            if ("next" in cls.lower()) or ("次" in text) or ("次" in aria) or page_vals:
                return abs_url, "HTMLから次ページURL検出"

    # 2) ページ番号リンクの中から、現在ページより大きい最小ページを探す
    page_candidates = []
    for a in soup.select('a[href*="page="]'):
        href = a.get("href", "")
        abs_url = make_absolute_url(href, current_url)
        qs = parse_qs(urlparse(abs_url).query)
        vals = qs.get("page")
        if not vals:
            continue
        try:
            n = int(vals[0])
        except Exception:
            continue
        if n > page_count:
            page_candidates.append((n, abs_url))
    if page_candidates:
        page_candidates.sort(key=lambda x: x[0])
        return page_candidates[0][1], "ページ番号リンクから次ページURL検出"

    # 3) 最後の手段：URLにpage=Nを手動で付ける
    fallback_url = set_query_param(current_url, "page", page_count + 1)
    return fallback_url, "手動でpageパラメータ生成"

# =============================================
#   メインのスクレイピング制御関数
# =============================================
def scrape_creema_fast(start_url, max_num, include_eval_dates=True, include_first_review_date=True):
    reset_diagnostics()
    diag_log("開始", f"リサーチ開始: 上限={max_num}件 / 速度={speed_mode} / リトライ={retry_count}回 / 待機={retry_base_wait}秒", start_url, "INFO")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "ja,en-US;q=0.9,en;q=0.8"
    }
    today = datetime.now()
    one_month_ago = today - timedelta(days=30)
    
    all_item_elements_data = []
    seen_urls = set()  # 取得済みURLの重複チェック用

    current_url = start_url
    page_count = 1
    detected_market_total = 0 
    page_status = st.empty()
    
    while current_url and len(all_item_elements_data) < max_num:
        page_status.info(
            f" ページ巡回中... 現在 {page_count} ページ目をスキャンしています "
            f"(収集済リンク: {len(all_item_elements_data)}件)"
        )

        try:
            response = fast_get(current_url, headers=headers, timeout=10, retry_label="一覧ページ")
            diag_count("一覧ページ_アクセス")
            diag_log("一覧ページ", f"{page_count}ページ目 status={response.status_code} / bytes={len(response.content)}", current_url, "INFO")

            if response.status_code != 200:
                diag_log("一覧ページ", f"status={response.status_code} のため一覧巡回を終了", current_url, "ERROR")
                break

            soup = BeautifulSoup(response.content, "html.parser")
            
            if page_count == 1:
                search_count_element = soup.find(string=re.compile(r"検索結果\s*[\d,]+件"))
                if search_count_element:
                    match_count = re.search(r"検索結果\s*([\d,]+)件", search_count_element)
                    if match_count:
                        detected_market_total = int(match_count.group(1).replace(",", ""))
            
            items = soup.select("article.c-item-article")

            if not items:
                diag_log("一覧ページ", "商品ブロック article.c-item-article が0件。URL違い・HTML変更・アクセス制限の可能性", current_url, "ERROR")
                break

            before_count = len(all_item_elements_data)
            duplicate_count = 0
                
            for item in items:
                if len(all_item_elements_data) >= max_num:
                    break

                title_tag = item.select_one('.c-item-article__name a[href*="/item/"]')

                if not title_tag:
                    continue
                    
                title = title_tag.text.strip()
                href = title_tag.get("href", "")

                if not href:
                    continue

                if href.startswith("http"):
                    raw_link = href
                else:
                    raw_link = "https://www.creema.jp" + href

                # vkey / screen / profile などのパラメータ違いを同じ商品として扱う
                link = normalize_creema_item_url(raw_link)

                # URL全文ではなく、正規化した商品URLで重複を防ぐ
                if link in seen_urls:
                    duplicate_count += 1
                    diag_count("一覧ページ_正規化URL重複スキップ")
                    continue

                seen_urls.add(link)
                
                desc_tag = item.select_one(".c-item-article__desc")
                creator, price = "取得失敗", 0

                if desc_tag and "/" in desc_tag.text:
                    parts = desc_tag.text.split("/")
                    price = int(re.sub(r"\D", "", parts[0])) if parts[0] else 0
                    creator = parts[1].strip()
                
                all_item_elements_data.append({
                    "link": link,
                    "creator": creator,
                    "title": title,
                    "price": price
                })
            
            added_count = len(all_item_elements_data) - before_count
            diag_count("一覧ページ_商品リンク追加", added_count)
            if duplicate_count:
                diag_count("一覧ページ_重複URLスキップ", duplicate_count)
            diag_log("一覧ページ", f"{page_count}ページ目: 商品候補 {len(items)}件 / 追加 {added_count}件 / 重複スキップ {duplicate_count}件 / 累計 {len(all_item_elements_data)}件", current_url, "INFO")

            # 次ページURL取得。HTML変更に強くするため、複数の方法で探す。
            if len(all_item_elements_data) >= max_num:
                current_url = None
            else:
                next_url, next_reason = find_next_listing_url(soup, current_url, page_count)

                # 手動生成URLの場合、本当に商品があるかは次ループで確認する。
                # 商品0件・同じ商品だけの場合は、その時点で終了する。
                if next_url and next_url != current_url:
                    diag_log("一覧ページ", f"次ページへ移動: {next_reason}", next_url, "INFO")
                    current_url = next_url
                    page_count += 1
                    time.sleep(CURRENT_SPEED["sleep"])
                else:
                    diag_log("一覧ページ", "次ページURLが作れないため一覧巡回を終了", current_url, "WARN")
                    current_url = None

            # 次ページへ進んでも新規追加が0件の状態が続くと無限巡回になり得るため、ここで止める。
            # 「全部重複だった」場合だけでなく、「HTML変更等でこのページから
            # 1件も商品を拾えなかった」場合も同様に止める（理由を問わず追加0件なら終了）。
            if added_count == 0 and len(items) > 0:
                diag_log("一覧ページ", f"このページから新規追加0件（重複{duplicate_count}件/候補{len(items)}件）のため一覧巡回を終了", current_url, "WARN")
                current_url = None

        except Exception as e:
            diag_log("一覧ページ", f"例外: {type(e).__name__}: {e}", current_url, "ERROR")
            print("一覧ページ取得エラー:", e)
            print(traceback.format_exc())
            break
            
    page_status.empty()
    total_found = len(all_item_elements_data)
    diag_log("一覧ページ", f"商品リンク収集完了: {total_found}件", start_url, "INFO")
    if total_found == 0:
        diag_log("終了", "商品リンクが0件のため終了。検索キーワード/URL/HTML変更/アクセス制限を確認してください", start_url, "ERROR")
        return {"items": [], "market_total": detected_market_total, "diagnostics": get_diagnostics()}
        
    status_text = st.empty()
    progress_bar = st.progress(0)
    scraped_data = []
    
    max_workers = CURRENT_SPEED["workers_large"] if total_found > 100 else CURRENT_SPEED["workers_small"]
    current_idx = 0
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_item = {
            executor.submit(_internal_fetch_item, item_data, headers, one_month_ago, include_eval_dates, include_first_review_date): item_data 
            for item_data in all_item_elements_data
        }
        for future in as_completed(future_to_item):
            item_data = future_to_item[future]
            try:
                result = future.result()
            except Exception as e:
                result = None
                diag_count("Future_例外")
                diag_log("詳細解析", f"future.result()で例外: {type(e).__name__}: {e}", item_data.get("link", ""), "ERROR")
                print(traceback.format_exc())
            current_idx += 1
            if result:
                scraped_data.append(result)
            else:
                diag_count("詳細解析_結果なし")
            progress_bar.progress(min(current_idx / total_found, 1.0))
            status_text.text(f"⏳ 解析中（{speed_mode}）... 完了: {current_idx} / {total_found} 件 / 成功: {len(scraped_data)} 件")
                
    progress_bar.empty()
    status_text.empty()
    
    diag_log("詳細解析", f"詳細解析完了: 成功 {len(scraped_data)}件 / 失敗 {total_found - len(scraped_data)}件", start_url, "INFO")
    if scraped_data:
        # 念のため、詳細解析後にも商品URLを正規化して重複除外する
        unique_data = []
        seen_result_urls = set()
        removed_after_detail = 0

        for item in scraped_data:
            normalized_url = normalize_creema_item_url(item.get("商品URL", ""))
            item["商品URL"] = normalized_url

            if normalized_url in seen_result_urls:
                removed_after_detail += 1
                continue

            seen_result_urls.add(normalized_url)
            unique_data.append(item)

        if removed_after_detail:
            diag_count("詳細解析後_重複除外", removed_after_detail)
            diag_log("詳細解析", f"詳細解析後に重複商品を {removed_after_detail} 件除外", "", "WARN")

        for i, item in enumerate(unique_data, 1):
            item["No."] = i

        return {"items": unique_data, "market_total": detected_market_total, "diagnostics": get_diagnostics()}
    diag_log("終了", "商品リンクは取れたが、詳細解析に成功した商品が0件でした", start_url, "ERROR")
    return {"items": [], "market_total": detected_market_total, "diagnostics": get_diagnostics()}

# =============================================
#    ⚙️ アプリ実行処理エリア
# =============================================
if "raw_data" not in st.session_state: st.session_state.raw_data = None
if "market_total" not in st.session_state: st.session_state.market_total = 0
if "diagnostics" not in st.session_state: st.session_state.diagnostics = None

if start_button:
    if (mode == "一覧URL直貼り" and not target_url) or (mode == "キーワード検索" and not search_keyword):
        st.error("⚠️ 条件を入力してください。")
    else:
        cond_text = f"キーワード: {search_keyword}" if mode == "キーワード検索" else f"直貼りURL: {target_url}"
        st.session_state.diagnostics = None

        if not skip_line_notify:
            try:
                send_line_notification(cond_text, max_items)
            except Exception as e:
                # LINE通知で止まらないようにする
                st.warning(f"LINE通知でエラーが出ましたが、リサーチは続行します: {type(e).__name__}: {e}")

        with st.spinner("🔄 Creemaのデータを解析中...（サーバー負荷を抑えるため、ゆっくり進みます）"):
            res_dict = scrape_creema_fast(target_url, max_items, include_eval_dates, include_first_review_date)

        if res_dict is not None:
            st.session_state.raw_data = res_dict.get("items", [])
            st.session_state.include_eval_dates = include_eval_dates
            st.session_state.include_first_review_date = include_first_review_date
            st.session_state.market_total = res_dict.get("market_total", 0)
            st.session_state.diagnostics = res_dict.get("diagnostics")

            if len(st.session_state.raw_data) > 0:
                st.success(f"🎉 リサーチ完了！ 全 {len(st.session_state.raw_data)} 件のデータを取得しました。")
            else:
                st.error("❌ データが0件でした。条件やCreema側の状態を確認してください。")

        else:
            st.error("❌ データが取得できませんでした。条件やCreema側の状態を確認してください。")
            st.session_state.diagnostics = get_diagnostics()

# --- 診断ログ表示（何が起きているかを見えるようにする） ---
if st.session_state.get("diagnostics"):
    logs, stats = st.session_state.diagnostics
    with st.expander(f"🩺 診断ログを見る（直近の集計: {dict(stats)}）", expanded=False):
        if logs:
            st.dataframe(pd.DataFrame(logs), use_container_width=True, hide_index=True)
        else:
            st.caption("ログはまだありません。")

# --- 画面表示処理 ---
if st.session_state.raw_data is not None and len(st.session_state.raw_data) > 0:
    raw_df = pd.DataFrame(st.session_state.raw_data)

    # 古い取得結果が session_state に残っている場合の保険
    raw_df = raw_df.rename(columns={
        "総評価数": "作家の総評価数",
        "直近販売日1": "評価日1",
        "直近販売日2": "評価日2",
        "直近販売日3": "評価日3",
        "一番初めの評価日": "作家の一番初めの評価日",
    })

    # 必要な列がない場合の保険
    for col in ["評価日1", "評価日2", "評価日3"]:
        if col not in raw_df.columns:
            raw_df[col] = "ー"

    if "作家の総評価数" not in raw_df.columns:
        raw_df["作家の総評価数"] = 0

    if "その商品の直近1ヶ月の評価数" not in raw_df.columns:
        raw_df["その商品の直近1ヶ月の評価数"] = "0件"

    # 数値変換の安全処理
    raw_df["価格(円)"] = pd.to_numeric(raw_df["価格(円)"], errors="coerce").fillna(0).astype(int)
    raw_df["お気に入り数"] = pd.to_numeric(raw_df["お気に入り数"], errors="coerce").fillna(0).astype(int)
    raw_df["作家の総評価数"] = pd.to_numeric(raw_df["作家の総評価数"], errors="coerce").fillna(0).astype(int)

    # 購入者数の数値化
    def parse_buyer_count(val):
        if not isinstance(val, str):
            return 0
        if "10人以上" in val:
            return 10
        match = re.search(r"(\d+)", val)
        return int(match.group(1)) if match else 0

    # その商品の直近1ヶ月の評価数の数値化
    def parse_recent_review_count(val):
        # 表示上は「42件」のような文字列で入ってくることがあるが、
        # 並び替えを正しくするため、内部では必ず数値として扱う。
        if pd.isna(val):
            return 0
        if isinstance(val, (int, float, np.integer, np.floating)):
            return int(val)
        match = re.search(r"(\d+)", str(val))
        return int(match.group(1)) if match else 0

    # 「その商品の直近1ヶ月の評価数」は文字列のままだと、42件が4件の近くに並ぶなど
    # 文字列順ソートになってしまうため、表示前に数値型へ変換する。
    raw_df["その商品の直近1ヶ月の評価数"] = raw_df["その商品の直近1ヶ月の評価数"].apply(parse_recent_review_count).astype(int)

    # 日付変換
    def parse_to_date(val):
        if not isinstance(val, str):
            return None
        match = re.search(r"(\d{4})\.(\d{2})\.(\d{2})", val)
        if match:
            return datetime.strptime(match.group(0), "%Y.%m.%d").date()
        return None


    def filter_row(row):
        # 価格
        if not (min_price <= row["価格(円)"] <= max_price):
            return False

        # その商品の直近1ヶ月の評価数
        recent_review_num = parse_recent_review_count(row.get("その商品の直近1ヶ月の評価数", "0件"))
        if not (min_recent_review <= recent_review_num <= max_recent_review):
            return False


        # 作家の総評価数
        if not (min_rev <= row["作家の総評価数"] <= max_rev):
            return False



        return True

    # フィルター適用
    mask = raw_df.apply(filter_row, axis=1)
    filtered_df = raw_df[mask].copy()

    if not filtered_df.empty:
        filtered_df["No."] = range(1, len(filtered_df) + 1)

    # 列順を整える
    preferred_columns = [
        "No.",
        "商品URL",
        "作家名",
        "商品名",
        "価格(円)",
        "作品紹介文",
        "その商品の直近1ヶ月の評価数",
        "作家の総評価数",
        "購入者数",
        "お気に入り数",
    ]

    if st.session_state.get("include_eval_dates", True):
        preferred_columns += ["評価日1", "評価日2", "評価日3"]
    else:
        filtered_df = filtered_df.drop(columns=["評価日1", "評価日2", "評価日3"], errors="ignore")

    if st.session_state.get("include_first_review_date", True):
        preferred_columns += ["作家の一番初めの評価日"]
    else:
        filtered_df = filtered_df.drop(columns=["作家の一番初めの評価日"], errors="ignore")

    existing_columns = [col for col in preferred_columns if col in filtered_df.columns]
    other_columns = [col for col in filtered_df.columns if col not in existing_columns]
    filtered_df = filtered_df[existing_columns + other_columns]

    st.markdown(f"**現在の表示件数:** {len(filtered_df)} 件 / 全体 {len(raw_df)} 件")

    st.dataframe(
        filtered_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "No.": st.column_config.NumberColumn(
                "No.",
                width="small"
            ),
            "商品URL": st.column_config.LinkColumn(
                "商品URL",
                display_text="商品ページ",
                width="small"
            ),
            "作家名": st.column_config.TextColumn(
                "作家名",
                width="small"
            ),
            "商品名": st.column_config.TextColumn(
                "商品名",
                width="small"
            ),
            "価格(円)": st.column_config.NumberColumn(
                "価格(円)",
                width="small"
            ),
            "作品紹介文": st.column_config.TextColumn(
                "作品紹介文",
                width="medium"
            ),
            "その商品の直近1ヶ月の評価数": st.column_config.NumberColumn(
                "その商品の直近1ヶ月の評価数",
                help="数値として並び替えできます",
                format="%d 件",
                width="small"
            ),
            "作家の総評価数": st.column_config.NumberColumn(
                "作家の総評価数",
                width="small"
            ),
            "購入者数": st.column_config.TextColumn(
                "購入者数",
                width="small"
            ),
            "お気に入り数": st.column_config.NumberColumn(
                "お気に入り数",
                width="small"
            ),
            "評価日1": st.column_config.TextColumn(
                "評価日1",
                width="small"
            ),
            "評価日2": st.column_config.TextColumn(
                "評価日2",
                width="small"
            ),
            "評価日3": st.column_config.TextColumn(
                "評価日3",
                width="small"
            ),
            "作家の一番初めの評価日": st.column_config.TextColumn(
                "作家の一番初めの評価日",
                width="small"
            ),
        }
    )

    excel_data = convert_df_to_excel(filtered_df)

    st.download_button(
        label="📥 絞り込んだデータをExcelでダウンロード",
        data=excel_data,
        file_name=f"creema_research_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
